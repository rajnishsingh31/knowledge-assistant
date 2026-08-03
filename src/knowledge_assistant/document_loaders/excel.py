from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from knowledge_assistant.document_loaders.base import DocumentLoader
from knowledge_assistant.document_loaders.common import (
    create_document,
    validate_file,
)
from knowledge_assistant.models import Document


class ExcelDocumentLoader(DocumentLoader):
    """Extract worksheet values from XLSX files."""

    @property
    def supported_extensions(self) -> frozenset[str]:
        return frozenset({".xlsx"})

    def load(self, path: Path) -> Document:
        resolved_path = validate_file(path)

        workbook = load_workbook(
            filename=resolved_path,
            read_only=True,
            data_only=True,
        )

        try:
            sheet_sections: list[str] = []

            for worksheet in workbook.worksheets:
                sheet_lines = [
                    f"[Worksheet: {worksheet.title}]"
                ]

                for row in worksheet.iter_rows(
                    values_only=True
                ):
                    values = [
                        self._format_cell(value)
                        for value in row
                    ]

                    while values and not values[-1]:
                        values.pop()

                    if any(values):
                        sheet_lines.append(
                            " | ".join(values)
                        )

                if len(sheet_lines) > 1:
                    sheet_sections.append(
                        "\n".join(sheet_lines)
                    )

            return create_document(
                path=resolved_path,
                content="\n\n".join(sheet_sections),
            )
        finally:
            workbook.close()

    @staticmethod
    def _format_cell(value: Any) -> str:
        if value is None:
            return ""

        return str(value).strip()